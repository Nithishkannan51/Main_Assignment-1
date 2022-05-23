import Admin

import openpyxl
from xlwt import Workbook
from xlrd import open_workbook as rd

from xlutils.copy import copy

import user

ad=Admin.admin()
us=user.user()

class Login():


    print("******Welcome to BookMyShow******* ")
    print("1.Login ")
    print("2.Register new account ")
    print("3.Exit ")
    print("Please enter an valid input as number")
    integer=int(input())
    username_data=["Nithish","Sidharth","Arka","Gupta","Kumar"]
    password_data=["nithish12","sidharth34","arka56","gupta78","kumar90"]

    count=0

    if(integer==1):
        print("******Welcome to BookMyShow******* ")
        username=input("Your username :")
        password=input("Your password :")
        print(username)
        print(password)
        if username=="admin123" and password=="admin":
            print("******Welcome Admin******* ")
            print("1. Add New Movie Info ")
            print("2. Edit Movie Info ")
            print("3. Delete Movies ")
            print("4.Logout")
            num = int(input())
            if(num==1):
                ad.addnewmovie()
            elif(num==2):
                ad.editmovie()
            elif(num==3):
                ad.Deletemovie()




        else:

            for i in range(len(username_data)):
                if username==username_data[i] and password==password_data[i]:

                    count=count+1

            if(count==0):
                print("Invalid username or password try again later")
            else:

                print("******Welcome "+username+" *******")

                r = rd("Movielist.xls")
                s = r.sheet_by_index(0)

                size_col = s.ncols
                size_row = s.nrows
                #wb1 = copy(r)
                #w_sheet = wb1.get_sheet(0)
                for i in range(1,size_col):
                    j=str(i)
                    x=str(s.cell(0, i).value)
                    print(j+". "+x)
                print("Logout ")

                num1 = int(input("Enter movie number: "))

                #Delete_movie = input("Select movie which you want to Delete:")
                for i in range(8):
                        print(s.cell(i,0).value + " : "+str(s.cell(i, num1).value))

                print("1.  Book Tickets")
                print("2.  Cancel Tickets")
                print("3.  Give User Rating ")
                num2=int(input())

                if(num2==1):
                    us.bookticket(num1)
                elif(num2==2):
                    us.cancelticket(num1)
                elif(num2==3):
                    us.userrating()
                else:
                    print("Invalid input try again now")
                    while (num2 != 1, 2, 3):
                        num2 = int(input("Enter input again "))

                        if (num2 == 1):
                            us.bookticket()
                            break
                        elif (num2 == 2):
                            us.cancelticket()
                            break
                        elif (num2 == 3):
                            us.userrating()
                            break
                        else:
                            print("Invalid input try again now")



    elif(integer==2):
        print("register new user")
        print("****Create new Account***** ")
        r = rd("Movielist.xls")
        s = r.sheet_by_index(1)

        size_col = s.ncols
        size_row = s.nrows
        wb1 = copy(r)
        w_sheet = wb1.get_sheet(1)

        newusername=input("Enter Name:")
        for i in range(1,size_row):
            if(newusername==s.cell(i, 0).value):
                print("Entered new user already exists ")
                newusername = input("Enter Name:")
                break
        newemail=input("Enter Email:")
        newphone=input("Enter phone:")
        newAge=input("Enter Age:")
        newPassword=input("Enter password:")
        list1=[newusername,newPassword,newemail,newphone,newAge]


        for i in range(5):
            w_sheet.write(size_row,i,list1[i])
        wb1.save('Movielist.xls')

    elif(integer==3):
        print("You are logged out from the application")







