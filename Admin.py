import openpyxl
from openpyxl import load_workbook
import xlutils
import xlwt
import xlrd


from xlwt import Workbook
from xlrd import open_workbook as rd

from xlutils.copy import copy

class admin():

    def __init__(self):
        self.title=""
        self.genre=""
        self.length=""
        self.cast=""
        self.director=""
        self.adminrating=""
        self.language=""
        self.timing=""
        self.numberofshows=""
        self.firstshow=""
        self.intervaltime=""
        self.gap=""
        self.capacity=""
        self.list1 = ["title", "genre", "length", "cast", "directors", "adminrating", "language", "timing", "numberofshows",
                "firstshow", "intervaltime", "gap", "capacity"]
        self.wb = Workbook()
        # add_sheet is used to create sheet.
        self.sheet1 = self.wb.add_sheet('Sheet 1')

    def addnewmovie(self):
        new_dict = {}

        title,genre,length,cast,directors,adminrating,language,timing,numberofshows,firstshow,intervaltime,gap,capacity=self.title,self.genre,self.length,self.cast,self.director,self.adminrating,self.language,self.timing,self.numberofshows,self.firstshow,self.intervaltime,self.gap,self.capacity

        r = rd("Movielist.xls")

        s = r.sheet_by_index(0)

        size_col = s.ncols
        size_row = s.nrows
        wb1 = copy(r)
        sheet1 = wb1.get_sheet(0)


        mi_max_row = s.nrows
        mi_max_col = s.ncols
        print(mi_max_col)
        print(mi_max_row)
        list=["title","genre","length","cast","directors","adminrating","language","timing","numberofshows","firstshow","intervaltime","gap","capacity"]
        global size
        size=int(input("Please tell the number of movies"))



        for i in range(mi_max_col,mi_max_col+size):
            for j in range(len(list)):
                new_dict[list[j]] = input("Please enter " + list[j] + " :")
                a=new_dict[list[j]]

                sheet1.write(j,i,a)
                if(i==1):
                    sheet1.write(j,i-1,list[j])

        #print(new_dict)
        wb1.save('Movielist.xls')



    def editmovie(self):

        r = rd("Movielist.xls")

        s = r.sheet_by_index(0)

        size = s.ncols

        wb1=copy(r)
        w_sheet=wb1.get_sheet(0)
        #print(s.max_column)
        movie_edit=input("Select movie which you want to edit: ")
        mi_max_col =size

        for i in range (1,mi_max_col):

            a = s.cell(0, i).value
            if(movie_edit==a):
                for j in range(len(self.list1)):
                    a1= input("Please enter " + self.list1[j] + " :")

                    w_sheet.write(j,i,a1)

        wb1.save('Movielist.xls')
        #print(a)


    def Deletemovie(self):
        r = rd("Movielist.xls")
        s = r.sheet_by_index(0)

        size_col = s.ncols
        size_row = s.nrows
        wb1 = copy(r)
        w_sheet = wb1.get_sheet(0)

        Delete_movie=input("Select movie which you want to Delete:")
        for i in range(size_col):
            if(Delete_movie==s.cell(0, i).value):
                print(s.cell(0, i).value)
                a=""
                for j in range(size_row):
                    w_sheet.write(j, i, a)


        wb1.save('Movielist.xls')


